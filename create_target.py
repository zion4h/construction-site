import os
import numpy as np
import pandas as pd
import pymysql
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

"""
INPUT
- names 工资表的构成工人
- moneys 工人对应工资
"""
names = ["彭名刚", "朱红兵"]
moneys = [25000, 25000]

pricePersonDay = 350
pricePerBox = 35

if len(names) != len(moneys):
    raise BaseException("输入人名和工资数额数量未对齐")

# 设置 pd 打印
pd.set_option('display.max_columns', None)  # 显示所有列
pd.set_option('display.max_rows', None)  # 显示所有行
pd.set_option('display.width', None)  # 不折行显示


def get_worker_info(name):
    """
    根据姓名查询工人信息
    """
    config = {
        "host": "localhost",
        "user": "root",
        "password": "root",
        "database": "工地"
    }
    db = pymysql.connect(**config)
    sql = "SELECT 序号, 姓名, 性别, 工种, 电话, 家庭住址, 身份证, 合同签订时间, 进场时间, 银行卡号, 建行开户行 FROM 工人 WHERE 姓名=%s"

    # 执行查询
    cursor = db.cursor()
    c = cursor.execute(sql, (name,))
    ret = cursor.fetchone()
    # 关闭数据库连接
    cursor.close()
    db.close()

    return ret


# 用列表保存查询结果
results = []
results_index = []
real_moneys = []
for i in range(len(names)):
    result = get_worker_info(names[i])
    if result:
        results.append(result)
        results_index.append(result[0])
        real_moneys.append(moneys[i])
    else:
        print("未找到工人【" + names[i] + "】")

# 将查询结果转化为DataFrame
columns = ['序号', '姓名', '性别', '工种', '电话', '家庭住址', '身份证', '合同签订时间', '进场时间', '银行卡号',
           '建行开户行']
df = pd.DataFrame(results, index=results_index, columns=columns).sort_index()
# print(df)

# 计算数据
df.loc[:, '本月应发工资'] = real_moneys
df.loc[:, '应领工资'] = real_moneys
df.loc[:, '实发工资'] = real_moneys
df.loc[:, '单价/量'] = pricePerBox
df.loc[:, '单价/天'] = pricePersonDay
df.loc[:, '加班工资'] = 0
# 记工
df.loc[:, '本月工作量'] = np.around(np.divide(real_moneys, pricePersonDay))
df.loc[:, '累计工作量'] = np.around(np.divide(real_moneys, pricePersonDay))
df.loc[:, '出勤天数'] = np.around(np.divide(real_moneys, pricePersonDay))
df['加班工资'] = np.subtract(real_moneys, np.multiply(df['本月工作量'], pricePersonDay))
# 按照指定属性顺序重新排列列
df[['完成工作量', '借支', '农民工签字', '工资卡类别', '金额', '备注', '个人确认签字', '领款签字', '离场时间', '扣个税']] = np.nan

# 1. 农民工花名册
columns_order_民工花名册 = ['序号', '姓名', '身份证', '性别', '工种', '电话', '家庭住址', '合同签订时间', '进场时间', '离场时间']
df1 = df[columns_order_民工花名册]
output_path_民工花名册 = "民工花名册.xlsx"
df1.to_excel(output_path_民工花名册, index=False)

print(f"生成 {output_path_民工花名册} 成功")

# 2. 农民工工资表
columns_order_民工工资表 = ['序号', '姓名', '身份证', '工种', '累计工作量', '本月工作量', '单价/天', '加班工资',
                              '本月应发工资', '实发工资', '农民工签字']
df2 = df[columns_order_民工工资表]
output_path_民工工资表 = "民工工资表.xlsx"
df2.to_excel(output_path_民工工资表, index=False)

print(f"生成 {output_path_民工工资表} 成功")

# 农民工工资代发明细表
columns_order_民工工资代发明细表 = ['序号', '姓名', '银行卡号', '工资卡类别', '身份证', '电话', '本月应发工资',
                                      '建行开户行', '备注']
df3 = df[columns_order_民工工资代发明细表]
output_path_民工工资代发明细表 = "民工工资代发明细表.xlsx"
df3.to_excel(output_path_民工工资代发明细表, index=False)

print(f"生成 {output_path_民工工资代发明细表} 成功")

# 项目工人工资确认发放表
columns_order_工资确认发放表 = ['序号', '姓名', '性别', '身份证', '电话', '银行卡号', '单价/天', '本月工作量',
                                        '单价/量', '完成工作量', '加班工资', '本月应发工资', '借支', '实发工资',
                                        '个人确认签字', '领款签字', '备注']
df4 = df[columns_order_工资确认发放表]
output_path_工资确认发放表 = "工资确认发放表.xlsx"
df4.to_excel(output_path_工资确认发放表, index=False)

print(f"生成 {output_path_工资确认发放表} 成功")

# 劳工工资表
isLabor = False
if isLabor:
    months = 4
    df['扣个税'] = np.subtract(df['应领工资'], 5000 * (months + 1)).multiply(0.03)
    df['扣个税'] = [i if i >= 0 else 0 for i in df['扣个税']]
    columns_order_劳工工资表 = ['序号', '姓名', '身份证', '出勤天数', '单价/天', '应领工资', '扣个税', '实发工资',
                                '个人确认签字']
    df5 = df[columns_order_劳工工资表]
    output_path_劳工工资表 = "劳工工资表.xlsx"
    df5.to_excel(output_path_劳工工资表, index=False)

    print(f"生成 {output_path_劳工工资表} 成功")


# 优化输出表格的格式
def reset_col(filename):
    if not os.path.exists(filename):
        print("sorry, %s is not exit" % filename)
        return

    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        excel_df = pd.read_excel(filename, sheet).fillna('-')
        excel_df.loc[len(excel_df)] = list(excel_df.columns)  # 把标题行附加到最后一行
        for col in excel_df.columns:
            index = list(excel_df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = excel_df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen + 1  # 也就是列宽为最大长度+1 可以自己调整

        alignment_center = Alignment(horizontal='center', vertical='center')

        # 指定区域单元格居中
        ws_area = ws["A1:J100"]
        for i in ws_area:
            for j in i:
                j.alignment = alignment_center

    wb.save(filename)


paths = ['民工花名册.xlsx',
         '民工工资表.xlsx',
         '民工工资代发明细表.xlsx',
         '工资确认发放表.xlsx',
         '劳工工资表.xlsx',
         '奇怪的测试文件']
for path in paths:
    reset_col(path)
